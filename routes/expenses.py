"""Travel expenses + reports endpoints.

Mounted at /api/* by app.py.
"""

import calendar
import logging
from datetime import date
from pathlib import Path
from uuid import uuid4

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse
from pydantic import BaseModel

from db import get_db, next_invoice_number
from helpers import convert_to_chf
from generate_invoice import generate_expense_report

log = logging.getLogger(__name__)

router = APIRouter(tags=["expenses"])

_paths = {}
_ctx = {}


def configure(scan_dir: Path, report_dir: Path, company: str,
              aed_to_chf: float, supported_ext: set, get_customer_fn,
              analyze_receipt_fn, compute_file_hash_fn, is_duplicate_scan_fn):
    _paths["SCAN_DIR"] = scan_dir
    _paths["REPORT_DIR"] = report_dir
    _ctx["COMPANY"] = company
    _ctx["AED_TO_CHF"] = aed_to_chf
    _ctx["SUPPORTED_EXT"] = supported_ext
    _ctx["get_customer"] = get_customer_fn
    _ctx["analyze_receipt"] = analyze_receipt_fn
    _ctx["compute_file_hash"] = compute_file_hash_fn
    _ctx["is_duplicate_scan"] = is_duplicate_scan_fn


class FolderImport(BaseModel):
    path: str


class BulkIds(BaseModel):
    ids: list[int]


class BulkRecategorize(BaseModel):
    ids: list[int]
    category: str


@router.get("/expenses")
async def list_expenses(year: int | None = None):
    with get_db() as db:
        if year:
            rows = db.execute(
                "SELECT * FROM expenses WHERE substr(expense_date,1,4)=? ORDER BY expense_date",
                (str(year),),
            ).fetchall()
        else:
            rows = db.execute(
                "SELECT * FROM expenses ORDER BY expense_date DESC"
            ).fetchall()
    keys = rows[0].keys() if rows else []
    return [
        {
            "id": r["id"],
            "expense_date": r["expense_date"],
            "description": r["description"],
            "amount": r["amount"],
            "category": r["category"],
            "original_amount": r["original_amount"],
            "original_currency": r["original_currency"],
            "scan_file": r["scan_file"],
            "has_scan": r["scan_file"] is not None,
            "scan_type": (r["scan_file"].rsplit(".", 1)[-1] if r["scan_file"] else None),
            "trip_id": (r["trip_id"] if "trip_id" in keys else None),
        }
        for r in rows
    ]


@router.post("/expenses")
async def create_expense(
    expense_date: str = Form(...),
    description: str = Form(...),
    amount: float = Form(...),
    category: str = Form(...),
    scan: UploadFile = File(None),
):
    SCAN_DIR = _paths["SCAN_DIR"]
    from helpers import normalize_image_bytes, hashed_filename
    scan_filename = None
    if scan and scan.filename:
        raw = await scan.read()
        ext, data = normalize_image_bytes(scan.filename, raw)
        scan_filename = hashed_filename("exp", ext, data)
        path = SCAN_DIR / scan_filename
        if not path.exists():
            path.write_bytes(data)

    with get_db() as db:
        cur = db.execute(
            "INSERT INTO expenses (expense_date, description, amount, category, scan_file) VALUES (?,?,?,?,?)",
            (expense_date, description, amount, category, scan_filename),
        )

    return {"id": cur.lastrowid}


@router.post("/expenses/import-folder")
async def import_folder(data: FolderImport):
    SCAN_DIR = _paths["SCAN_DIR"]
    SUPPORTED_EXT = _ctx["SUPPORTED_EXT"]
    analyze_receipt = _ctx["analyze_receipt"]
    compute_file_hash = _ctx["compute_file_hash"]
    is_duplicate_scan = _ctx["is_duplicate_scan"]

    folder = Path(data.path).expanduser().resolve()
    if not folder.is_dir():
        raise HTTPException(400, f"Folder not found: {data.path}")

    import llm
    if not llm.status()["reachable"]:
        raise HTTPException(400, f"LLM provider '{llm.LLM_PROVIDER}' not reachable. "
                                  f"Check OLLAMA_URL or ANTHROPIC_API_KEY.")

    files = sorted(
        f for f in folder.iterdir()
        if f.suffix.lower() in SUPPORTED_EXT and not f.name.startswith(".")
    )
    if not files:
        raise HTTPException(400, "No supported images found (JPG/PNG)")

    results = []
    duplicates = 0
    for img_path in files:
        try:
            file_hash = compute_file_hash(img_path)
            if is_duplicate_scan(file_hash):
                duplicates += 1
                results.append({
                    "file": img_path.name,
                    "status": "ok",
                    "duplicate": True,
                    "date": "-",
                    "description": "Duplicate - skipped",
                    "amount": 0,
                    "category": "-",
                })
                continue

            data_extracted = analyze_receipt(img_path)

            ext = img_path.suffix.lower()
            scan_filename = f"exp_{uuid4().hex[:10]}{ext}"
            (SCAN_DIR / scan_filename).write_bytes(img_path.read_bytes())

            expense_date = data_extracted["date"]
            description = data_extracted["description"]
            original_amount = float(data_extracted["amount"])
            currency = data_extracted.get("currency", "AED").upper()
            category = data_extracted["category"]

            amount_chf = convert_to_chf(original_amount, currency)

            orig_amt = original_amount if currency != "CHF" else None
            orig_cur = currency if currency != "CHF" else None

            with get_db() as db:
                cur = db.execute(
                    """INSERT INTO expenses
                       (expense_date, description, amount, category, original_amount, original_currency, scan_file)
                       VALUES (?,?,?,?,?,?,?)""",
                    (expense_date, description, amount_chf, category, orig_amt, orig_cur, scan_filename),
                )

            results.append({
                "id": cur.lastrowid,
                "file": img_path.name,
                "date": expense_date,
                "description": description,
                "amount": amount_chf,
                "category": category,
                "status": "ok",
                "duplicate": False,
            })
        except Exception as e:
            log.exception("Failed to process %s", img_path.name)
            results.append({
                "file": img_path.name,
                "status": "error",
                "error": str(e),
            })

    ok = sum(1 for r in results if r["status"] == "ok" and not r.get("duplicate"))
    return {"imported": ok, "total": len(results), "duplicates": duplicates, "results": results}


@router.get("/expenses/years")
async def expense_years():
    with get_db() as db:
        rows = db.execute(
            "SELECT DISTINCT substr(expense_date,1,4) as y FROM expenses ORDER BY y"
        ).fetchall()
    return [r["y"] for r in rows]


@router.get("/expenses/summary")
async def expense_summary():
    with get_db() as db:
        rows = db.execute("""
            SELECT substr(expense_date,1,4) as year,
                   COUNT(*) as count,
                   SUM(amount) as total
            FROM expenses GROUP BY year ORDER BY year
        """).fetchall()
    return [{"year": r["year"], "count": r["count"], "total": r["total"]} for r in rows]


@router.get("/expenses/reports")
async def list_reports():
    with get_db() as db:
        rows = db.execute(
            "SELECT * FROM expense_reports ORDER BY year DESC, "
            "COALESCE(month, 0) DESC, id DESC"
        ).fetchall()
    return [
        {
            "id": r["id"],
            "report_number": r["report_number"],
            "year": r["year"],
            "month": r["month"] if "month" in r.keys() else None,
            "total": r["total"],
            "total_chf": r["total"],
            "expense_count": r["expense_count"],
            "created_at": r["created_at"],
        }
        for r in rows
    ]


def _report_filename(year: int, month: int | None, rpt_num: int) -> str:
    if month:
        return f"expenses_{year}_{month:02d}_{rpt_num:04d}.pdf"
    return f"expenses_{year}_{rpt_num:04d}.pdf"


@router.delete("/expenses/reports/{id}")
async def delete_report(id: int):
    REPORT_DIR = _paths["REPORT_DIR"]
    with get_db() as db:
        row = db.execute("SELECT * FROM expense_reports WHERE id=?", (id,)).fetchone()
        if not row:
            raise HTTPException(404, "Report not found")

        rpt_num = row["report_number"]
        year = row["year"]
        month = row["month"] if "month" in row.keys() else None

        pdf_path = REPORT_DIR / _report_filename(year, month, rpt_num)
        if pdf_path.exists():
            pdf_path.unlink()

        excel_path = REPORT_DIR / f"expenses_{year}.xlsx"
        if excel_path.exists() and not month:
            excel_path.unlink()

        db.execute("DELETE FROM invoices WHERE invoice_number=?", (rpt_num,))
        db.execute("DELETE FROM expense_reports WHERE id=?", (id,))

    return {"message": f"Report #{rpt_num:04d} deleted"}


@router.post("/expenses/report/{year}")
async def generate_report(year: int, month: int | None = None):
    SCAN_DIR = _paths["SCAN_DIR"]
    REPORT_DIR = _paths["REPORT_DIR"]
    get_customer = _ctx["get_customer"]

    with get_db() as db:
        if month:
            rows = db.execute(
                "SELECT * FROM expenses "
                "WHERE substr(expense_date,1,4)=? AND substr(expense_date,6,2)=? "
                "ORDER BY expense_date",
                (str(year), f"{month:02d}"),
            ).fetchall()
            scope = f"{year}-{month:02d}"
        else:
            rows = db.execute(
                "SELECT * FROM expenses WHERE substr(expense_date,1,4)=? "
                "ORDER BY expense_date",
                (str(year),),
            ).fetchall()
            scope = str(year)

    if not rows:
        raise HTTPException(400, f"No expenses found for {scope}")

    # Replace any existing report for this exact (year, month) scope — but
    # KEEP its report number (it is the invoice number the accountant and the
    # client already hold). Only a brand-new scope takes the next number.
    keep_number = None
    keep_reimbursed = None
    keep_total = None
    with get_db() as db:
        if month:
            existing = db.execute(
                "SELECT report_number FROM expense_reports WHERE year=? AND month=?",
                (year, month),
            ).fetchall()
        else:
            existing = db.execute(
                "SELECT report_number FROM expense_reports WHERE year=? AND month IS NULL",
                (year,),
            ).fetchall()
        for old in existing:
            if keep_number is None:
                keep_number = old["report_number"]
                prev = db.execute(
                    "SELECT reimbursed_at, total FROM expense_reports WHERE report_number=?",
                    (old["report_number"],)).fetchone()
                keep_reimbursed = prev["reimbursed_at"] if prev else None
                keep_total = float(prev["total"] or 0) if prev else None
            old_pdf = REPORT_DIR / _report_filename(year, month, old["report_number"])
            if old_pdf.exists():
                old_pdf.unlink()
            db.execute("DELETE FROM invoices WHERE invoice_number=?",
                       (old["report_number"],))
        if month:
            db.execute("DELETE FROM expense_reports WHERE year=? AND month=?",
                       (year, month))
        else:
            db.execute("DELETE FROM expense_reports WHERE year=? AND month IS NULL",
                       (year,))
        # only purge the yearly excel when regenerating the year-wide report
        if not month:
            old_excel = REPORT_DIR / f"expenses_{year}.xlsx"
            if old_excel.exists():
                old_excel.unlink()

    expenses = [
        {
            "date": r["expense_date"],
            "description": r["description"],
            "category": r["category"],
            "amount": r["amount"],
            "original_amount": r["original_amount"],
            "original_currency": r["original_currency"],
            "scan_path": str(SCAN_DIR / r["scan_file"]) if r["scan_file"] else None,
        }
        for r in rows
    ]

    inv_num = keep_number if keep_number is not None else next_invoice_number()
    total = sum(e["amount"] for e in expenses)

    with get_db() as db:
        customer = get_customer(db, 1)

    pdf_bytes = generate_expense_report(year, expenses, inv_num,
                                         customer=customer, month=month)
    report_path = REPORT_DIR / _report_filename(year, month, inv_num)
    report_path.write_bytes(pdf_bytes)

    MONTH_NAMES = ["January","February","March","April","May","June",
                   "July","August","September","October","November","December"]
    period_label = f"{MONTH_NAMES[month-1]} {year}" if month else f"{year}"

    with get_db() as db:
        issued = date.today()
        due_m = issued.month + 1 if issued.month < 12 else 1
        due_y = issued.year if issued.month < 12 else issued.year + 1
        due_date = date(due_y, due_m, calendar.monthrange(due_y, due_m)[1])
        # Invoice "month" column reflects either the trip month or year-end.
        inv_month_col = month if month else 12
        db.execute(
            """INSERT INTO invoices
               (invoice_number, year, month, hours, rate, vat_rate,
                subtotal, tax, total, issued_date, due_date, notes)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (inv_num, year, inv_month_col, 0, 0, 0,
             total, 0, total, str(issued), str(due_date),
             f"Travel expenses {period_label}"),
        )
        # A regenerated report that was already reimbursed stays reimbursed
        # only if its total didn't change; otherwise it reopens so the delta
        # shows up in the Reimburse dialog.
        if keep_reimbursed and keep_total is not None and abs(keep_total - total) > 0.005:
            keep_reimbursed = None
        db.execute(
            "INSERT INTO expense_reports "
            "(report_number, year, month, total, expense_count, reimbursed_at) VALUES (?,?,?,?,?,?)",
            (inv_num, year, month, total, len(expenses), keep_reimbursed),
        )

    return {"report_number": inv_num, "year": year, "month": month,
            "total": total, "count": len(expenses), "regenerated": keep_number is not None}


@router.get("/expenses/report/{year}/pdf")
async def download_report(year: int, download: bool = False, month: int | None = None):
    REPORT_DIR = _paths["REPORT_DIR"]
    COMPANY = _ctx["COMPANY"]
    with get_db() as db:
        if month:
            row = db.execute(
                "SELECT * FROM expense_reports WHERE year=? AND month=? "
                "ORDER BY id DESC LIMIT 1",
                (year, month),
            ).fetchone()
        else:
            row = db.execute(
                "SELECT * FROM expense_reports WHERE year=? AND month IS NULL "
                "ORDER BY id DESC LIMIT 1",
                (year,),
            ).fetchone()
    if not row:
        raise HTTPException(404, "No report for this scope")

    rpt_num = row["report_number"]
    path = REPORT_DIR / _report_filename(year, month, rpt_num)
    if not path.exists():
        raise HTTPException(404, "Report PDF not found")

    MONTH_NAMES = ["January","February","March","April","May","June",
                   "July","August","September","October","November","December"]
    period_label = f"{MONTH_NAMES[month-1]} {year}" if month else f"{year}"

    if download:
        return FileResponse(
            path,
            filename=f"Travel Expenses {period_label} {COMPANY} 101119.LOD-SW_GCS-24032.pdf",
            media_type="application/pdf",
        )
    return FileResponse(path, media_type="application/pdf")


@router.get("/expenses/report/{year}/excel")
async def download_excel(year: int, month: int | None = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    import io

    REPORT_DIR = _paths["REPORT_DIR"]
    COMPANY = _ctx["COMPANY"]
    AED_TO_CHF = _ctx["AED_TO_CHF"]

    MONTH_NAMES = ["January","February","March","April","May","June",
                   "July","August","September","October","November","December"]
    period_label = f"{MONTH_NAMES[month-1]} {year}" if month else f"{year}"
    file_stem = f"expenses_{year}_{month:02d}" if month else f"expenses_{year}"

    with get_db() as db:
        if month:
            rows = db.execute(
                "SELECT * FROM expenses "
                "WHERE substr(expense_date,1,4)=? AND substr(expense_date,6,2)=? "
                "ORDER BY expense_date",
                (str(year), f"{month:02d}"),
            ).fetchall()
        else:
            rows = db.execute(
                "SELECT * FROM expenses WHERE substr(expense_date,1,4)=? "
                "ORDER BY expense_date",
                (str(year),),
            ).fetchall()
    if not rows:
        raise HTTPException(404, f"No expenses for {period_label}")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Expenses {period_label}"[:31]  # Excel sheet name max 31 chars

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="D7E1E8")
    chf_fmt = '#,##0.00'
    thin_border = Border(bottom=Side(style="thin", color="E2E8F0"))

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Muster Consulting GmbH - Travel Expenses {period_label}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Exchange rate: 1 AED = {AED_TO_CHF} CHF"
    ws["A2"].font = Font(size=9, color="888888")

    headers = ["Date", "Description", "Category", "Amount (CHF)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for i, r in enumerate(rows, 5):
        ws.cell(row=i, column=1, value=r["expense_date"])
        ws.cell(row=i, column=2, value=r["description"])
        ws.cell(row=i, column=3, value=r["category"])
        amt_cell = ws.cell(row=i, column=4, value=r["amount"])
        amt_cell.number_format = chf_fmt
        for col in range(1, 5):
            ws.cell(row=i, column=col).border = thin_border

    total_row = len(rows) + 5
    ws.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True)
    total_val = sum(r["amount"] for r in rows)
    total_cell = ws.cell(row=total_row, column=4, value=total_val)
    total_cell.number_format = chf_fmt
    total_cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    excel_path = REPORT_DIR / f"{file_stem}.xlsx"
    excel_path.write_bytes(buf.getvalue())

    return FileResponse(
        excel_path,
        filename=f"Travel Expenses {period_label} {COMPANY} 101119.LOD-SW_GCS-24032.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.post("/expenses/bulk/delete")
async def bulk_delete_expenses(data: BulkIds):
    SCAN_DIR = _paths["SCAN_DIR"]
    with get_db() as db:
        for eid in data.ids:
            row = db.execute("SELECT scan_file FROM expenses WHERE id = ?", (eid,)).fetchone()
            if row and row["scan_file"]:
                scan_path = SCAN_DIR / row["scan_file"]
                if scan_path.exists():
                    scan_path.unlink()
            db.execute("DELETE FROM expenses WHERE id = ?", (eid,))
    return {"deleted": len(data.ids)}


@router.post("/expenses/bulk/recategorize")
async def bulk_recategorize_expenses(data: BulkRecategorize):
    valid_cats = {"Meals", "Transport", "Accommodation", "Other"}
    if data.category not in valid_cats:
        raise HTTPException(400, f"Invalid category: {data.category}")
    with get_db() as db:
        for eid in data.ids:
            db.execute("UPDATE expenses SET category=? WHERE id=?", (data.category, eid))
    return {"updated": len(data.ids)}


@router.get("/expenses/{id}")
async def get_expense(id: int):
    with get_db() as db:
        row = db.execute("SELECT * FROM expenses WHERE id = ?", (id,)).fetchone()
    if not row:
        raise HTTPException(404, "Expense not found")
    return {
        "id": row["id"],
        "expense_date": row["expense_date"],
        "description": row["description"],
        "amount": row["amount"],
        "category": row["category"],
        "scan_file": row["scan_file"],
        "has_scan": row["scan_file"] is not None,
    }


@router.put("/expenses/{id}")
async def update_expense(
    id: int,
    expense_date: str = Form(...),
    description: str = Form(...),
    amount: float = Form(...),
    category: str = Form(...),
    scan: UploadFile = File(None),
):
    SCAN_DIR = _paths["SCAN_DIR"]
    with get_db() as db:
        row = db.execute("SELECT * FROM expenses WHERE id = ?", (id,)).fetchone()
        if not row:
            raise HTTPException(404, "Expense not found")

        scan_filename = row["scan_file"]
        if scan and scan.filename:
            from helpers import normalize_image_bytes, hashed_filename
            raw = await scan.read()
            ext, data = normalize_image_bytes(scan.filename, raw)
            new_name = hashed_filename("exp", ext, data)
            # Only delete the old scan if it isn't reused by other rows AND the
            # content hash changed (otherwise we'd delete the file we're about to
            # reuse, or break a sibling row referencing the same blob).
            if scan_filename and scan_filename != new_name:
                still_used = db.execute(
                    "SELECT 1 FROM expenses WHERE scan_file=? AND id!=? LIMIT 1",
                    (scan_filename, id),
                ).fetchone()
                if not still_used:
                    old = SCAN_DIR / scan_filename
                    if old.exists():
                        old.unlink()
            scan_filename = new_name
            path = SCAN_DIR / scan_filename
            if not path.exists():
                path.write_bytes(data)

        db.execute(
            """UPDATE expenses SET expense_date=?, description=?, amount=?,
               category=?, scan_file=? WHERE id=?""",
            (expense_date, description, amount, category, scan_filename, id),
        )

    return {"message": "Expense updated"}


@router.delete("/expenses/{id}")
async def delete_expense(id: int):
    SCAN_DIR = _paths["SCAN_DIR"]
    with get_db() as db:
        row = db.execute("SELECT scan_file FROM expenses WHERE id = ?", (id,)).fetchone()
        if not row:
            raise HTTPException(404, "Expense not found")
        scan_file = row["scan_file"]
        db.execute("DELETE FROM expenses WHERE id = ?", (id,))
        # Only unlink the shared blob if no other row references it (content-hash
        # filenames mean the same scan can be referenced by multiple expenses).
        if scan_file:
            still_used = db.execute(
                "SELECT 1 FROM expenses WHERE scan_file=? LIMIT 1", (scan_file,),
            ).fetchone()
            if not still_used:
                scan_path = SCAN_DIR / scan_file
                if scan_path.exists():
                    scan_path.unlink()

    return {"message": "Expense deleted"}


@router.get("/expenses/{id}/scan")
async def get_scan(id: int):
    SCAN_DIR = _paths["SCAN_DIR"]
    with get_db() as db:
        row = db.execute("SELECT scan_file FROM expenses WHERE id = ?", (id,)).fetchone()
    if not row or not row["scan_file"]:
        raise HTTPException(404, "No scan found")
    scan_path = SCAN_DIR / row["scan_file"]
    if not scan_path.exists():
        raise HTTPException(404, "Scan file not found")
    return FileResponse(scan_path)
