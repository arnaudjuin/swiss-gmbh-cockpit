"""Bank statement Excel export (multi-sheet workbook).

Extracted from routes/bank.py — same URL, same behavior. Classification
rules are computed inside the endpoint using the shared payroll helpers,
mirroring the on-screen Owner-ledger logic.
Mounted at /api/* by app.py.
"""

from fastapi import APIRouter, HTTPException

from db import get_db
from routes.bank import list_transactions

router = APIRouter(tags=["bank-export"])

@router.get("/bank-statements/{id}/export.xlsx")
async def export_transactions_xlsx(
    id: int,
    quarter: int | None = None,
    year: int | None = None,
):
    """Rich, multi-sheet Excel export for a bank statement.

    Optional filtering:
      ?quarter=1|2|3|4&year=YYYY → only transactions falling in that quarter

    Sheets:
      1. Summary — Kontokorrent recap, cash flow split, reimbursement matches
      2. Transactions — flat rows with Classification; parent aggregators
         ("multi e-banking order") skipped; sub-entries expanded in place
      3. Reimbursements — 1 row per (bank inflow → expense report) match

    All classification (salary / reimbursement / personal / intra-company /
    routine payroll) is computed server-side using the same rules as the UI.
    """
    from datetime import datetime as _dt
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io
    import re as _re

    # Fetch parsed transactions. id == 0 → ALL statements merged into one
    # full-history export (same classification, lifetime Kontokorrent).
    _merged_stmt = None
    if id == 0:
        data, _merged_stmt = await _merge_all_statements()
    else:
        data = await list_transactions(id)
    if isinstance(data, dict) and "error" in data:
        raise HTTPException(400, data["error"])
    txs = data.get("transactions", []) or []
    # Keep an unfiltered copy so account-side detection stays consistent
    # regardless of quarter filter (a Q1 slice may not have any payroll yet).
    all_txs = txs
    currency = data.get("currency") or "CHF"

    # ── Optional quarter filter ──────────────────────────────────────────
    period_label_suffix = ""
    quarter_start = quarter_end = None
    if quarter and quarter in (1, 2, 3, 4):
        # Default year: from the statement period if not passed
        if not year:
            ps = data.get("period_start") or ""
            year = int(ps[:4]) if len(ps) >= 4 and ps[:4].isdigit() else _dt.today().year
        q_first_month = 3 * (quarter - 1) + 1
        import calendar as _cal
        last_day = _cal.monthrange(year, q_first_month + 2)[1]
        quarter_start = f"{year}-{q_first_month:02d}-01"
        quarter_end = f"{year}-{q_first_month + 2:02d}-{last_day:02d}"
        # Filter transactions AND their sub_entries
        def in_range(dstr):
            return dstr and quarter_start <= dstr[:10] <= quarter_end
        filtered = []
        for tx in txs:
            keep_tx = in_range(tx.get("date"))
            kept_subs = [s for s in (tx.get("sub_entries") or []) if in_range(s.get("date") or tx.get("date"))]
            if keep_tx or kept_subs:
                tx2 = {**tx}
                if not keep_tx:
                    # Parent's own amount is out of range but some subs are in
                    # → keep only the surviving subs, hide the parent aggregator amount
                    tx2["amount"] = 0
                tx2["sub_entries"] = kept_subs if tx2.get("sub_entries") else []
                filtered.append(tx2)
        txs = filtered
        # Recompute totals in/out/net on the filtered set
        total_in_calc = 0.0
        total_out_calc = 0.0
        for tx in txs:
            if not tx.get("sub_entries"):
                amt = float(tx.get("amount") or 0)
                if amt > 0: total_in_calc += amt
                elif amt < 0: total_out_calc += amt
            else:
                for s in tx["sub_entries"]:
                    amt = float(s.get("amount") or 0)
                    if amt > 0: total_in_calc += amt
                    elif amt < 0: total_out_calc += amt
        data = dict(data)  # avoid mutating cached response
        data["total_in"] = round(total_in_calc, 2)
        data["total_out"] = round(total_out_calc, 2)
        data["net"] = round(total_in_calc + total_out_calc, 2)
        data["period_start"] = quarter_start
        data["period_end"] = quarter_end
        period_label_suffix = f" — Q{quarter} {year}"

    # Fetch statement meta from DB
    with get_db() as db:
        if _merged_stmt is not None:
            stmt_d = _merged_stmt
        else:
            stmt = db.execute(
                "SELECT * FROM bank_statements WHERE id=?", (id,)
            ).fetchone()
            if not stmt:
                raise HTTPException(404, "Statement not found")
            stmt_d = dict(stmt)
        # Expense reports for reimbursement matching
        reports = [dict(r) for r in db.execute(
            "SELECT id, report_number, year, month, total, expense_count, created_at "
            "FROM expense_reports ORDER BY report_number"
        ).fetchall()]
        # Payroll settings for salary detection + counterparty patterns
        payroll_row = db.execute(
            "SELECT * FROM payroll_settings ORDER BY id DESC LIMIT 1"
        ).fetchone()
        # Company expenses paid with the owner's personal card in the effective
        # period — they never hit this bank account but belong in the recap
        # (Kontokorrent: GmbH owes them back to the owner).
        eff_start = quarter_start or stmt_d.get("period_start") or "0000-01-01"
        eff_end = quarter_end or stmt_d.get("period_end") or "9999-12-31"
        personal_card_rows = [dict(pc) for pc in db.execute(
            """SELECT doc_date, vendor, description, category, amount, currency, status,
                      doc_file, reimbursed_at
               FROM company_docs WHERE paid_via='personal' AND doc_date >= ? AND doc_date <= ?
               ORDER BY doc_date""",
            (eff_start, eff_end),
        ).fetchall()]
        # Reimbursement transfers logged via "Reimburse yourself" — their bank
        # payments must not be misread as new non-salary debt to the owner.
        reimb_transfers = [dict(r) for r in db.execute(
            """SELECT transfer_date, amount FROM account_transfers
               WHERE direction='gmbh_to_personal'
               AND description LIKE 'Personal-card reimbursement%'""",
        ).fetchall()]
        # Owner contributions logged in the ledger — bank credits matching
        # these are owner money-in (GmbH owes it back), not revenue.
        owner_in_transfers = [dict(r) for r in db.execute(
            """SELECT transfer_date, amount FROM account_transfers
               WHERE direction='personal_to_gmbh'""",
        ).fetchall()]
    # Period-correct Kontokorrent: a bill still counts as debt in this window
    # if it was unreimbursed at the window's end. (Its settling bank payment
    # falls in the later window, where it is excluded — see classifier below.)
    personal_card_total = round(sum(
        float(pc["amount"] or 0) for pc in personal_card_rows
        if not pc["reimbursed_at"] or pc["reimbursed_at"] > eff_end), 2)
    # Travel expense reports the owner fronted (same Kontokorrent logic:
    # counts while unreimbursed at the window's end; legacy ones are stamped)
    with get_db() as db:
        er_rows = [dict(r) for r in db.execute(
            "SELECT report_number, total, created_at, reimbursed_at FROM expense_reports "
            "WHERE substr(created_at,1,10) >= ? AND substr(created_at,1,10) <= ?",
            (eff_start, eff_end)).fetchall()]
    expense_reports_total = round(sum(
        float(r["total"] or 0) for r in er_rows
        if not r["reimbursed_at"] or (r["reimbursed_at"] != "legacy" and r["reimbursed_at"] > eff_end)), 2)

    # ── Classification helpers ───────────────────────────────────────────
    # Reuse the shared payroll computation so net-salary matches everywhere.
    from routes.payroll import _row_to_settings, _compute_payslip
    monthly_net = 0.0
    employee_name = ""
    employer_name = "Muster Consulting"
    pay_day = 25
    if payroll_row:
        settings = _row_to_settings(payroll_row)
        calc = _compute_payslip(settings)
        monthly_net = float(calc.get("net_salary") or 0)
        employee_name = (settings.get("employee_name") or "").strip().lower()
        employer_name = (settings.get("employer_name") or "Muster Consulting").strip().lower()
        pay_day = int(settings.get("payment_day") or 25)
    emp_tokens = [t for t in employee_name.split() if len(t) >= 4]
    first_emp_token = emp_tokens[0] if emp_tokens else ""

    def matches_employee(cp: str) -> bool:
        if not emp_tokens: return False
        c = (cp or "").lower()
        if first_emp_token and first_emp_token not in c: return False
        return sum(1 for t in emp_tokens if t in c) >= 2

    def matches_relative(cp: str) -> bool:
        """Shares ≥2 surname/middle tokens with the employee name but does NOT
        include the employee's first name → likely a family member (spouse,
        parent, sibling). Not the employee themselves, not a business client."""
        if not emp_tokens or not first_emp_token: return False
        c = (cp or "").lower()
        if first_emp_token in c: return False   # this is the employee, not a relative
        return sum(1 for t in emp_tokens if t in c) >= 2

    def matches_employer(cp: str) -> bool:
        return bool(employer_name) and employer_name in (cp or "").lower()

    def is_routine_payroll(tx: dict) -> bool:
        hay = f"{tx.get('counterparty','')} {tx.get('description','')}".lower()
        pats = [
            r"\b(salaire|salary|salaer|lohn|gehalt|wage|paie|payroll)",
            r"\b(quellensteuer|source.?tax|withhold|imp[oô]t.{0,8}source)",
            r"\b(ahv|avs|alv|aho|apg|caf|cas)\b",
            r"\b(bvg|lpp|pension|retirement|pr[eé]voyance|pilier|s[aä]ule)",
            r"\b(uvg|laa|suva|krankentag)",
            r"\b(3a|pillar.?3|pilier.?3|s[aä]ule.?3)\b",
            r"\b(vat|tva|mwst|iva)\b",
        ]
        return any(_re.search(p, hay) for p in pats)

    def parse_d(s):
        if not s: return None
        try: return _dt.strptime(s[:10], "%Y-%m-%d")
        except Exception: return None

    # Detect side of account — use the FULL statement, not the filtered slice,
    # so a quarter with no employee payments (e.g. Q1 before employment start)
    # doesn't misdetect this as a personal account.
    emp_hits = sum(1 for tx in all_txs if matches_employee(tx.get("counterparty", "")))
    empr_hits = sum(1 for tx in all_txs if matches_employer(tx.get("counterparty", "")))
    account_side = "gmbh" if emp_hits > empr_hits else "personal"

    # Salary detection
    # Candidate salary amounts: the current settings net PLUS, for each ledger
    # date that carries a 'Net salary' transfer, the SUM of that day's
    # GmbH→Personal transfers. A retroactive salary change splits an old
    # payment into salary + Kontokorrent repayment in the ledger while the
    # bank shows one line for the original amount — the per-day sum matches it.
    with get_db() as db:
        _sal_dates = [r["transfer_date"] for r in db.execute(
            "SELECT DISTINCT transfer_date FROM account_transfers WHERE description LIKE 'Net salary%'")]
        salary_candidates = []       # (ledger_date, day_total, salary_portion)
        for _d in _sal_dates:
            tot = db.execute(
                "SELECT COALESCE(SUM(amount),0) t FROM account_transfers "
                "WHERE transfer_date=? AND direction='gmbh_to_personal'", (_d,)).fetchone()
            sal = db.execute(
                "SELECT COALESCE(SUM(amount),0) t FROM account_transfers "
                "WHERE transfer_date=? AND direction='gmbh_to_personal' "
                "AND description LIKE 'Net salary%'", (_d,)).fetchone()
            salary_candidates.append((_d, round(float(tot["t"]), 2), round(float(sal["t"]), 2)))
    if monthly_net:
        salary_candidates.append((None, round(monthly_net, 2), round(monthly_net, 2)))

    def salary_candidate_for(tx: dict):
        """Matched (day_total, salary_portion) or None. Amount within ±10%,
        and for ledger-dated candidates the bank line must fall within ±7
        days of that ledger date; the settings-net candidate (dateless) uses
        the payday window instead."""
        amt = abs(float(tx.get("amount") or 0))
        d = parse_d(tx.get("date"))
        if not d:
            return None
        for led_date, total, sal in salary_candidates:
            if abs(amt - total) > max(100, total * 0.10):
                continue
            if led_date:
                ld = parse_d(led_date)
                if ld and abs((d - ld).days) <= 7:
                    return (total, sal)
            else:
                day_diff = abs(d.day - pay_day)
                if day_diff <= 7 or day_diff >= 23:
                    return (total, sal)
        return None

    def looks_like_salary(tx: dict) -> bool:
        return salary_candidate_for(tx) is not None

    # Reimbursement matching (positive inflows only, exact amount, after report creation)
    used_reports = set()
    reim_matches = []  # list of {"tx": ..., "report": ...}
    for tx in txs:
        for sub in tx.get("sub_entries", []) or []:
            pass  # sub-entries handled below when flattening
    flat_inflows = []
    for tx in txs:
        if (tx.get("amount") or 0) > 0 and not tx.get("sub_entries"):
            flat_inflows.append(tx)
        for sub in tx.get("sub_entries", []) or []:
            if (sub.get("amount") or 0) > 0:
                flat_inflows.append({**sub, "date": tx.get("date"), "transaction_no": tx.get("transaction_no")})
    for tx in flat_inflows:
        tx_d = parse_d(tx.get("date"))
        for r in reports:
            if r["id"] in used_reports: continue
            if abs((r["total"] or 0) - (tx.get("amount") or 0)) > 0.05: continue
            created = parse_d((r["created_at"] or "")[:10])
            if created and tx_d and created > tx_d: continue
            used_reports.add(r["id"])
            reim_matches.append({"tx": tx, "report": r})
            break

    def sig(tx):
        return f"{tx.get('date','')}|{float(tx.get('amount') or 0):.2f}|{(tx.get('counterparty','') or '').lower().strip()}"

    reim_labels = {}
    for m in reim_matches:
        r = m["report"]
        period = f"{r['year']}" + (f"-{r['month']:02d}" if r["month"] else "")
        reim_labels[sig(m["tx"])] = f"Reimbursement (report #{r['report_number']}, {period}, {r['expense_count']} receipts)"

    def is_pc_reimbursement(tx: dict) -> bool:
        """Outflow to the employee matching a logged personal-card
        reimbursement transfer (amount ±0.05, date within 10 days)."""
        amt = abs(float(tx.get("amount") or 0))
        d = parse_d(tx.get("date"))
        if not d:
            return False
        for t in reimb_transfers:
            td = parse_d(t["transfer_date"])
            if td and abs(float(t["amount"]) - amt) <= 0.05 and abs((d - td).days) <= 10:
                return True
        return False

    _owner_in_consumed = set()

    def is_owner_contribution(tx: dict) -> bool:
        """Credit matching a logged Personal → GmbH transfer (amount ±0.05,
        date within 10 days). Each logged row matches one transaction. A
        credit already matched to an expense report is the client's
        reimbursement — never double-counted here."""
        if sig(tx) in reim_labels:
            return False
        amt = float(tx.get("amount") or 0)
        d = parse_d(tx.get("date"))
        if amt <= 0 or not d:
            return False
        for idx, t in enumerate(owner_in_transfers):
            if idx in _owner_in_consumed:
                continue
            td = parse_d(t["transfer_date"])
            if td and abs(float(t["amount"]) - amt) <= 0.05 and abs((d - td).days) <= 10:
                _owner_in_consumed.add(idx)
                return True
        return False

    salary_sigs = set()
    pc_reimb_sigs = set()
    non_salary_paid = 0.0
    salary_paid = 0.0
    for tx in txs:
        if matches_employee(tx.get("counterparty", "")):
            if looks_like_salary(tx):
                salary_sigs.add(sig(tx))
                cand = salary_candidate_for(tx)
                amt = abs(float(tx.get("amount") or 0))
                part = min(amt, cand[1]) if cand else amt
                salary_paid += part
                non_salary_paid += amt - part   # retro-reclassified remainder
            elif float(tx.get("amount") or 0) < 0 and is_pc_reimbursement(tx):
                # Settles fronted personal-card bills — not new Kontokorrent debt
                pc_reimb_sigs.add(sig(tx))
            else:
                non_salary_paid += abs(float(tx.get("amount") or 0))

    reim_total = sum(float(m["tx"].get("amount") or 0) for m in reim_matches)
    total_in = float(data.get("total_in") or 0)
    total_out = float(data.get("total_out") or 0)
    net = float(data.get("net") or 0)

    # Intra-company inflows on a GmbH account: money that appears as coming
    # from "Muster Consulting" on Muster Consulting' own account = Sperrkonto release,
    # founding capital, inter-account transfer. NOT customer revenue.
    intra_in = 0.0
    if account_side == "gmbh":
        for tx in txs:
            amt = float(tx.get("amount") or 0)
            if amt > 0 and matches_employer(tx.get("counterparty", "")) and not matches_employee(tx.get("counterparty", "")):
                intra_in += amt
            for sub in tx.get("sub_entries") or []:
                samt = float(sub.get("amount") or 0)
                if samt > 0 and matches_employer(sub.get("counterparty", "")) and not matches_employee(sub.get("counterparty", "")):
                    intra_in += samt

    # Personal / family inflows: counterparty shares the employee's surname
    # (or middle names) but NOT the first name → likely a family member.
    # Excluded from "customer revenue" too. Credits matching a LOGGED
    # Personal → GmbH transfer are owner contributions instead (Kontokorrent).
    personal_family_in = 0.0
    owner_in = 0.0
    owner_in_sigs = set()
    for tx in txs:
        amt = float(tx.get("amount") or 0)
        if amt > 0 and (matches_relative(tx.get("counterparty", ""))
                        or matches_employee(tx.get("counterparty", ""))):
            if is_owner_contribution(tx):
                owner_in += amt
                owner_in_sigs.add(sig(tx))
            elif matches_relative(tx.get("counterparty", "")):
                personal_family_in += amt
        for sub in tx.get("sub_entries") or []:
            samt = float(sub.get("amount") or 0)
            if samt > 0 and matches_relative(sub.get("counterparty", "")):
                personal_family_in += samt
    if account_side == "gmbh" and reim_total > 0:
        residual = non_salary_paid - reim_total   # negative → GmbH owes
    else:
        # For pure non-salary case, present in absolute terms
        residual = -non_salary_paid if account_side == "gmbh" else non_salary_paid
    if account_side == "gmbh":
        # Expenses the owner fronted on a personal card → GmbH owes them back
        residual -= personal_card_total
        # Travel expense reports fronted privately, not yet paid back
        residual -= expense_reports_total
        # Money the owner put in from private accounts (logged in the ledger)
        residual -= owner_in
    if residual < 0:
        direction_text = f"GmbH owes you CHF {abs(residual):,.2f}"
    elif residual > 0:
        direction_text = f"You owe GmbH CHF {residual:,.2f}"
    else:
        direction_text = "Fully settled"

    def classify(tx: dict) -> str:
        if sig(tx) in reim_labels: return reim_labels[sig(tx)]
        if sig(tx) in salary_sigs: return "Salary"
        if sig(tx) in pc_reimb_sigs: return "Personal-card reimbursement (settles fronted bills)"
        if sig(tx) in owner_in_sigs: return "Owner contribution (logged in ledger)"
        cp = tx.get("counterparty", "")
        if matches_employee(cp):  return "Personal transfer (non-salary)"
        if matches_relative(cp):  return "Personal / family transfer"
        if matches_employer(cp):
            return "Intra-company transfer" if account_side == "gmbh" else "From/to employer"
        if is_routine_payroll(tx): return "Payroll / social charges"
        return ""

    # ── Build workbook ───────────────────────────────────────────────────
    wb = Workbook()

    HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
    HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
    SUB_FONT = Font(bold=True, size=10, color="1F3A5F")
    SUB_FILL = PatternFill("solid", fgColor="D7E1E8")
    ALT_FILL = PatternFill("solid", fgColor="F5F7FA")
    TITLE_FONT = Font(bold=True, size=14, color="1F3A5F")
    THIN = Side(style="thin", color="E2E8F0")
    BORDER = Border(bottom=THIN)
    CHF_FMT = '#,##0.00;[Red]-#,##0.00'
    DATE_FMT = 'yyyy-mm-dd'

    # ── Sheet 1: Summary ─────────────────────────────────────────────────
    s1 = wb.active
    s1.title = "Summary"

    def _label_val(ws, r, label, value, *, is_currency=False, bold=False, color=None):
        ws.cell(row=r, column=1, value=label).font = Font(bold=bold)
        c = ws.cell(row=r, column=2, value=value)
        if is_currency: c.number_format = CHF_FMT
        if bold: c.font = Font(bold=True)
        if color: c.font = Font(bold=bold, color=color)
        c.alignment = Alignment(horizontal="right")

    s1.merge_cells("A1:D1")
    s1["A1"] = f"Bank statement export — {stmt_d.get('bank','')} {stmt_d.get('account_label','')}{period_label_suffix}"
    s1["A1"].font = TITLE_FONT

    r = 3
    s1.cell(row=r, column=1, value="Period").font = Font(bold=True)
    # Show effective (filtered) period, not the statement's raw period
    effective_start = quarter_start or stmt_d.get('period_start', '')
    effective_end = quarter_end or stmt_d.get('period_end', '')
    s1.cell(row=r, column=2, value=f"{effective_start} → {effective_end}")
    r += 1
    s1.cell(row=r, column=1, value="IBAN").font = Font(bold=True)
    s1.cell(row=r, column=2, value=stmt_d.get("iban", ""))
    r += 1
    _label_val(s1, r, "Opening balance", stmt_d.get("opening_balance") or 0, is_currency=True); r += 1
    _label_val(s1, r, "Closing balance", stmt_d.get("closing_balance") or 0, is_currency=True, bold=True); r += 1

    r += 1
    s1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    s1.cell(row=r, column=1, value="Cash-flow summary").font = SUB_FONT
    s1.cell(row=r, column=1).fill = SUB_FILL
    r += 1
    _label_val(s1, r, "Total in", total_in, is_currency=True, color="16A34A"); r += 1
    _label_val(s1, r, "  of which customer revenue",
               total_in - reim_total - intra_in - personal_family_in - owner_in, is_currency=True); r += 1
    if owner_in > 0:
        _label_val(s1, r, "  of which owner contributions (logged in ledger)", owner_in, is_currency=True, color="9333EA"); r += 1
    _label_val(s1, r, "  of which travel reimbursement (excluded from revenue)", reim_total, is_currency=True, color="3B82F6"); r += 1
    if intra_in > 0:
        _label_val(s1, r, "  of which intra-company transfer (Sperrkonto / capital)", intra_in, is_currency=True, color="808080"); r += 1
    if personal_family_in > 0:
        _label_val(s1, r, "  of which personal / family transfer", personal_family_in, is_currency=True, color="808080"); r += 1
    _label_val(s1, r, "Total out", total_out, is_currency=True, color="DC2626"); r += 1
    _label_val(s1, r, "Net", net, is_currency=True, bold=True); r += 1
    if personal_card_rows:
        _label_val(s1, r, "Off-bank: company expenses paid by personal card (see 'Personal card' sheet)",
                   -personal_card_total, is_currency=True, color="9333EA"); r += 1
        _label_val(s1, r, "Total costs incl. personal card", total_out - personal_card_total,
                   is_currency=True, bold=True); r += 1

    r += 1
    s1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    s1.cell(row=r, column=1, value="Kontokorrent (GmbH ↔ Personal)").font = SUB_FONT
    s1.cell(row=r, column=1).fill = SUB_FILL
    r += 1
    _label_val(s1, r, "Non-salary paid to you", non_salary_paid, is_currency=True); r += 1
    _label_val(s1, r, "Reimbursement receivable (GmbH holds for you)", reim_total, is_currency=True, color="3B82F6"); r += 1
    if personal_card_rows:
        _label_val(s1, r, "Company expenses you paid with your personal card", personal_card_total,
                   is_currency=True, color="9333EA"); r += 1
    if expense_reports_total > 0:
        _label_val(s1, r, "Travel expense reports you fronted (not yet reimbursed)", expense_reports_total,
                   is_currency=True, color="9333EA"); r += 1
    if owner_in > 0:
        _label_val(s1, r, "Owner contributions from your private accounts", owner_in,
                   is_currency=True, color="9333EA"); r += 1
    _label_val(s1, r, "Residual balance", abs(residual), is_currency=True, bold=True, color="DC2626" if residual < 0 else "16A34A"); r += 1
    _label_val(s1, r, "Direction", direction_text, bold=True, color="DC2626" if residual < 0 else ("16A34A" if residual > 0 else "808080")); r += 1
    _label_val(s1, r, "Salary payments detected", len([s for s in salary_sigs])); r += 1
    _label_val(s1, r, "Total salary paid", salary_paid, is_currency=True); r += 1

    r += 1
    s1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    s1.cell(row=r, column=1, value="Reimbursement matches").font = SUB_FONT
    s1.cell(row=r, column=1).fill = SUB_FILL
    r += 1
    headers = ["Bank date", "Amount", "Counterparty", "Matched to report"]
    for i, h in enumerate(headers, 1):
        cell = s1.cell(row=r, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    r += 1
    for m in reim_matches:
        tx = m["tx"]; rep = m["report"]
        period = f"{rep['year']}" + (f"-{rep['month']:02d}" if rep['month'] else "")
        s1.cell(row=r, column=1, value=tx.get("date","")).alignment = Alignment(horizontal="left")
        c = s1.cell(row=r, column=2, value=float(tx.get("amount") or 0))
        c.number_format = CHF_FMT; c.alignment = Alignment(horizontal="right")
        c.font = Font(color="3B82F6", bold=True)
        s1.cell(row=r, column=3, value=tx.get("counterparty",""))
        s1.cell(row=r, column=4, value=f"#{rep['report_number']} ({period}, {rep['expense_count']} receipts)")
        r += 1

    # Column widths on Summary
    s1.column_dimensions["A"].width = 52
    s1.column_dimensions["B"].width = 18
    s1.column_dimensions["C"].width = 32
    s1.column_dimensions["D"].width = 40

    # ── Sheet 2: Transactions ────────────────────────────────────────────
    s2 = wb.create_sheet("Transactions")
    tx_headers = ["Date", "Value Date", "Amount", "Currency", "Counterparty",
                  "Description", "Transaction No.", "Classification", "Balance"]
    for i, h in enumerate(tx_headers, 1):
        cell = s2.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    s2.freeze_panes = "A2"
    s2.auto_filter.ref = f"A1:{get_column_letter(len(tx_headers))}1"

    # Zero-amount "Balance closing of service prices" rows are UBS statement
    # period markers, not real transactions. Skip them so the sheet only
    # contains meaningful movements.
    def _is_noise(tx: dict) -> bool:
        amt = float(tx.get("amount") or 0)
        desc = (tx.get("description") or "").lower()
        cpty = (tx.get("counterparty") or "").lower()
        return amt == 0 and ("balance closing" in desc or "balance closing" in cpty)

    row_idx = 2
    for tx in txs:
        has_subs = bool(tx.get("sub_entries"))
        if not has_subs:
            if _is_noise(tx):
                continue
            _write_tx_row(s2, row_idx, tx, currency, classify, CHF_FMT, ALT_FILL, is_sub=False, parent_tx_no=None)
            row_idx += 1
        else:
            # Emit only sub-entries, skip parent aggregator
            for sub in tx.get("sub_entries", []):
                effective = {
                    "date": tx.get("date"),
                    "value_date": tx.get("value_date"),
                    "amount": sub.get("amount"),
                    "counterparty": sub.get("counterparty"),
                    "description": sub.get("description"),
                    "transaction_no": tx.get("transaction_no"),
                    "balance": None,
                }
                if _is_noise(effective):
                    continue
                _write_tx_row(s2, row_idx, effective, currency, classify, CHF_FMT, ALT_FILL, is_sub=True,
                              parent_tx_no=tx.get("transaction_no"))
                row_idx += 1

    # Column widths on Transactions (9 columns now — Reference column dropped)
    widths = {"A": 12, "B": 12, "C": 14, "D": 10, "E": 34, "F": 42, "G": 18, "H": 46, "I": 14}
    for col, w in widths.items():
        s2.column_dimensions[col].width = w

    # ── Sheet 3: Reimbursements (drill-down) ─────────────────────────────
    s3 = wb.create_sheet("Reimbursements")
    reim_headers = ["Bank date", "Amount received", "Counterparty",
                    "Report number", "Report year/month", "Receipts", "Report created"]
    for i, h in enumerate(reim_headers, 1):
        cell = s3.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    s3.freeze_panes = "A2"

    for i, m in enumerate(reim_matches, 2):
        tx = m["tx"]; rep = m["report"]
        period = f"{rep['year']}" + (f"-{rep['month']:02d}" if rep['month'] else "")
        s3.cell(row=i, column=1, value=tx.get("date",""))
        c = s3.cell(row=i, column=2, value=float(tx.get("amount") or 0))
        c.number_format = CHF_FMT; c.font = Font(color="3B82F6", bold=True)
        s3.cell(row=i, column=3, value=tx.get("counterparty",""))
        s3.cell(row=i, column=4, value=f"#{rep['report_number']}")
        s3.cell(row=i, column=5, value=period)
        s3.cell(row=i, column=6, value=rep["expense_count"])
        s3.cell(row=i, column=7, value=(rep.get("created_at") or "")[:10])

    if reim_matches:
        total_row = len(reim_matches) + 2
        s3.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
        tc = s3.cell(row=total_row, column=2, value=reim_total)
        tc.number_format = CHF_FMT; tc.font = Font(bold=True, color="3B82F6")

    widths3 = {"A": 12, "B": 16, "C": 30, "D": 12, "E": 14, "F": 10, "G": 14}
    for col, w in widths3.items():
        s3.column_dimensions[col].width = w

    # ── Sheet 4: Personal card (company expenses fronted by the owner) ───
    # These never appear in the bank account movements; they are listed
    # separately but included in the Summary totals and Kontokorrent.
    s4 = wb.create_sheet("Personal card")
    pc_headers = ["Date", "Vendor", "Description", "Category", "Amount",
                  "Currency", "Status", "Receipt on file", "Reimbursed"]
    for i, h in enumerate(pc_headers, 1):
        cell = s4.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    s4.freeze_panes = "A2"
    s4.auto_filter.ref = f"A1:{get_column_letter(len(pc_headers))}1"

    for i, pc in enumerate(personal_card_rows, 2):
        s4.cell(row=i, column=1, value=pc["doc_date"])
        s4.cell(row=i, column=2, value=pc["vendor"])
        s4.cell(row=i, column=3, value=pc["description"])
        s4.cell(row=i, column=4, value=pc["category"])
        c = s4.cell(row=i, column=5, value=float(pc["amount"] or 0))
        c.number_format = CHF_FMT
        c.font = Font(color="9333EA")
        s4.cell(row=i, column=6, value=pc["currency"])
        s4.cell(row=i, column=7, value=pc["status"])
        s4.cell(row=i, column=8, value="yes" if pc["doc_file"] else "no")
        s4.cell(row=i, column=9, value=pc["reimbursed_at"] or "outstanding")
        if pc["reimbursed_at"]:
            s4.cell(row=i, column=9).font = Font(color="16A34A")
        if i % 2 == 0:
            for col in range(1, len(pc_headers) + 1):
                s4.cell(row=i, column=col).fill = ALT_FILL

    if personal_card_rows:
        total_row = len(personal_card_rows) + 2
        s4.cell(row=total_row, column=1, value="Still owed to you at period end").font = Font(bold=True)
        tc = s4.cell(row=total_row, column=5, value=personal_card_total)
        tc.number_format = CHF_FMT
        tc.font = Font(bold=True, color="9333EA")
    else:
        s4.cell(row=2, column=1, value="No company expenses paid by personal card in this period.")

    widths4 = {"A": 12, "B": 26, "C": 42, "D": 22, "E": 14, "F": 10, "G": 10, "H": 14, "I": 14}
    for col, w in widths4.items():
        s4.column_dimensions[col].width = w

    # ── Serialize ─────────────────────────────────────────────────────────
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    if quarter:
        filename = f"bank_transactions_Q{quarter}_{year}.xlsx"
    else:
        prefix = "bank_transactions_ALL_" if id == 0 else "bank_transactions_"
        filename = f"{prefix}{stmt_d.get('period_start','start')}_to_{stmt_d.get('period_end','end')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _salary_month_label(date_str):
    """'Salary MM/YY' from a payment date. Salary is paid ~25th of the month
    of service, so payments on day 1-7 belong to the previous month (late
    payments), everything else belongs to the payment's own month."""
    from datetime import datetime as _dt
    try:
        d = _dt.strptime((date_str or "")[:10], "%Y-%m-%d")
    except Exception:
        return "Salary"
    if d.day <= 7:
        m = d.month - 1 if d.month > 1 else 12
        y = d.year if d.month > 1 else d.year - 1
    else:
        m, y = d.month, d.year
    return f"Salary {m:02d}/{y % 100:02d}"


def _write_tx_row(ws, row, tx, currency, classify, chf_fmt, alt_fill, *, is_sub, parent_tx_no):
    """Write one transaction row across 9 columns:
       1=Date, 2=Value Date, 3=Amount, 4=Currency, 5=Counterparty,
       6=Description, 7=Transaction No., 8=Classification, 9=Balance."""
    from openpyxl.styles import Font, Alignment, PatternFill
    ws.cell(row=row, column=1, value=tx.get("date","")).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=2, value=tx.get("value_date","")).alignment = Alignment(horizontal="left")
    amt = float(tx.get("amount") or 0)
    c = ws.cell(row=row, column=3, value=amt)
    c.number_format = chf_fmt
    if amt > 0:
        c.font = Font(color="16A34A", bold=True)
    elif amt < 0:
        c.font = Font(color="DC2626")
    ws.cell(row=row, column=4, value=currency).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=5, value=tx.get("counterparty",""))
    ws.cell(row=row, column=7, value=parent_tx_no if is_sub else tx.get("transaction_no",""))
    cls = classify(tx) if not is_sub else classify({"date": tx.get("date"), "amount": amt,
                                                     "counterparty": tx.get("counterparty",""),
                                                     "description": tx.get("description","")})
    # Override description so it reflects what the tool detected, not the
    # payer-entered "Reason for payment" (which is often stale — the sender
    # frequently reuses the same "Salary 04/26" label for every transfer).
    raw_desc = (tx.get("description","") or "").strip()
    if cls == "Salary":
        display_desc = _salary_month_label(tx.get("date", ""))
    elif cls.startswith("Personal transfer") and raw_desc.lower().startswith("salary"):
        # Preserve the original label so the payer's intent stays visible,
        # but mark it as suspect since the tool disagreed with the payer.
        display_desc = f"(labeled '{raw_desc}' — not detected as salary)"
    else:
        display_desc = raw_desc
    ws.cell(row=row, column=6, value=display_desc)
    cell_cls = ws.cell(row=row, column=8, value=cls)
    if cls.startswith("Reimbursement"):
        cell_cls.font = Font(color="3B82F6", italic=True, bold=True)
    elif cls == "Salary":
        cell_cls.font = Font(color="808080", italic=True)
    elif cls.startswith("Personal transfer"):
        cell_cls.font = Font(color="F59E0B", italic=True)
    elif cls == "Personal / family transfer":
        cell_cls.font = Font(color="F59E0B", italic=True)
    elif cls == "Intra-company transfer":
        cell_cls.font = Font(color="808080", italic=True)
    elif cls:
        cell_cls.font = Font(color="808080", italic=True)
    bal = tx.get("balance")
    if bal is not None:
        b = ws.cell(row=row, column=9, value=float(bal))
        b.number_format = chf_fmt
    # zebra striping
    if is_sub or row % 2 == 0:
        for col in range(1, 10):
            if not ws.cell(row=row, column=col).fill.fgColor.rgb or ws.cell(row=row, column=col).fill.fgColor.rgb == "00000000":
                ws.cell(row=row, column=col).fill = alt_fill



async def _merge_all_statements():
    """Union of every statement with a machine-readable file, ordered by
    period. Returns (data, synthetic_stmt_dict) shaped like the single-
    statement path expects."""
    with get_db() as db:
        stmts = [dict(r) for r in db.execute(
            "SELECT * FROM bank_statements WHERE statement_file_xml IS NOT NULL "
            "ORDER BY period_start"
        ).fetchall()]
    if not stmts:
        raise HTTPException(400, "No machine-readable statements to export")
    merged_txs = []
    for st in stmts:
        d = await list_transactions(st["id"])
        if isinstance(d, dict) and "error" not in d:
            merged_txs.extend(d.get("transactions", []) or [])
    merged_txs.sort(key=lambda t: (t.get("date") or ""), reverse=True)
    total_in = total_out = 0.0
    for tx in merged_txs:
        rows = tx.get("sub_entries") or [tx]
        for t in rows:
            amt = float(t.get("amount") or 0)
            if amt > 0: total_in += amt
            elif amt < 0: total_out += amt
    first, last = stmts[0], stmts[-1]
    data = {
        "period_start": first["period_start"], "period_end": last["period_end"],
        "currency": last.get("currency") or "CHF",
        "total_in": round(total_in, 2), "total_out": round(total_out, 2),
        "net": round(total_in + total_out, 2),
        "transactions": merged_txs,
    }
    stmt_d = {
        "bank": last.get("bank") or "UBS",
        "account_label": last.get("account_label") or "",
        "iban": last.get("iban") or "",
        "period_start": first["period_start"], "period_end": last["period_end"],
        "opening_balance": first.get("opening_balance"),
        "closing_balance": last.get("closing_balance"),
        "notes": f"Full history — {len(stmts)} statements combined",
    }
    return data, stmt_d
