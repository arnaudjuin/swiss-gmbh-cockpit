"""Quarterly summary + Accountant Package endpoints.

Mounted at /api/reports/* by app.py.
"""

import calendar
import csv
import io
import zipfile
from datetime import date
from pathlib import Path

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse

from db import get_db

router = APIRouter(prefix="/reports", tags=["reports"])
tax_router = APIRouter(tags=["tax"])  # for /vat and /tax (no prefix)

_paths = {}
_ctx = {}


def _table_exists(db, table_name: str) -> bool:
    try:
        db.execute(f"SELECT 1 FROM {table_name} LIMIT 1")
        return True
    except Exception:
        return False

def configure(pdf_dir: Path, acct_dir: Path, report_dir: Path,
              obligation_types: dict, salary: float,
              scan_dir: Path | None = None, payslip_dir: Path | None = None,
              bank_dir: Path | None = None):
    _paths["PDF_DIR"] = pdf_dir
    _paths["ACCT_DIR"] = acct_dir
    _paths["REPORT_DIR"] = report_dir
    if scan_dir is not None:
        _paths["SCAN_DIR"] = scan_dir
    if payslip_dir is not None:
        _paths["PAYSLIP_DIR"] = payslip_dir
    if bank_dir is not None:
        _paths["BANK_DIR"] = bank_dir
    _ctx["OBLIGATION_TYPES"] = obligation_types
    _ctx["SALARY"] = salary


# Re-export OBLIGATION_TYPES and SALARY references via a property-like helper
class _Const:
    def __getattr__(self, key):
        return _ctx.get(key)
const = _Const()

OBLIGATION_TYPES_REF = lambda: _ctx["OBLIGATION_TYPES"]
SALARY_REF = lambda: _ctx["SALARY"]


@router.get("/quarterly/{year}/{quarter}")
async def quarterly_report(year: int, quarter: int):
    """Quarterly summary for AHV filings and general bookkeeping."""
    if quarter < 1 or quarter > 4:
        raise HTTPException(400, "Quarter 1-4")
    start_month = (quarter - 1) * 3 + 1
    end_month = start_month + 2
    months = list(range(start_month, end_month + 1))

    with get_db() as db:
        # Invoices issued this quarter
        inv_rows = db.execute(
            "SELECT * FROM invoices WHERE year=? AND month BETWEEN ? AND ? AND hours>0 ORDER BY month",
            (year, start_month, end_month),
        ).fetchall()
        # Salary: gross wage sum of payslips actually ISSUED this quarter
        # (the AHV-relevant Lohnsumme), with their exact AHV amounts — not a
        # 3× projection of current settings.
        ps = db.execute(
            """SELECT COALESCE(SUM(gross),0) AS gross, COUNT(*) AS n,
                      COALESCE(SUM(emp_ahv),0) AS emp_ahv,
                      COALESCE(SUM(employer_ahv),0) AS employer_ahv
               FROM payslips WHERE year=? AND month BETWEEN ? AND ?""",
            (year, start_month, end_month),
        ).fetchone()
        salary_quarter = ps["gross"]
        payslip_count = ps["n"]

        # Bills this quarter
        bill_rows = db.execute(
            """SELECT category, COUNT(*) as cnt, SUM(amount) as total
               FROM company_docs
               WHERE substr(doc_date,1,4)=? AND CAST(substr(doc_date,6,2) AS INTEGER) BETWEEN ? AND ?
               GROUP BY category ORDER BY total DESC""",
            (str(year), start_month, end_month),
        ).fetchall()

        # Obligations paid this quarter
        ob_rows = db.execute(
            """SELECT obligation_type, COUNT(*) as cnt, SUM(amount) as total
               FROM obligations
               WHERE substr(due_date,1,4)=? AND CAST(substr(due_date,6,2) AS INTEGER) BETWEEN ? AND ?
               GROUP BY obligation_type""",
            (str(year), start_month, end_month),
        ).fetchall()

    inv_total = sum(r["total"] for r in inv_rows)
    gross_income = inv_total + salary_quarter
    # Exact AHV amounts from the issued payslips (rates shown for reference)
    ahv_employee = ps["emp_ahv"]
    ahv_employer = ps["employer_ahv"]

    return {
        "year": year,
        "quarter": quarter,
        "period_label": f"Q{quarter} {year}",
        "months": months,
        "invoices": {
            "count": len(inv_rows),
            "total": inv_total,
            "items": [{"invoice_number": r["invoice_number"], "month": r["month"],
                       "total": r["total"], "hours": r["hours"]} for r in inv_rows],
        },
        "salary": {
            "monthly": SALARY_REF(),
            "quarterly_total": salary_quarter,
            "payslip_count": payslip_count,
        },
        "ahv_estimate": {
            "employee_rate_pct": 5.3,
            "employer_rate_pct": 5.3,
            "employee_contribution": round(ahv_employee, 2),
            "employer_contribution": round(ahv_employer, 2),
            "total": round(ahv_employee + ahv_employer, 2),
            "basis": f"exact amounts from {payslip_count} issued payslip(s)",
        },
        "gross_income": gross_income,
        "bills_by_category": [
            {"category": r["category"], "count": r["cnt"], "total": r["total"]}
            for r in bill_rows
        ],
        "bills_total": sum(r["total"] for r in bill_rows),
        "obligations_by_type": [
            {"type": OBLIGATION_TYPES_REF().get(r["obligation_type"], r["obligation_type"]),
             "count": r["cnt"], "total": r["total"]}
            for r in ob_rows
        ],
        "obligations_total": sum(r["total"] for r in ob_rows),
    }


@router.get("/accountant-package/{year}")
async def accountant_package(year: int):
    """One-click export: invoice PDFs + expense report + all accounting docs + P&L Excel."""
    import io
    import zipfile
    import csv

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        with get_db() as db:
            # Invoices (exclude reimbursement reports — they're in expenses/, not income)
            invoices = db.execute(
                "SELECT * FROM invoices WHERE year=? AND hours>0 ORDER BY month",
                (year,),
            ).fetchall()
            inv_csv = io.StringIO()
            w = csv.writer(inv_csv)
            w.writerow(["Invoice #", "Month", "Hours", "Subtotal", "VAT", "Total", "Status", "Paid Date", "Due Date"])
            for r in invoices:
                w.writerow([
                    f"{r['invoice_number']:04d}", calendar.month_name[r["month"]],
                    r["hours"], r["subtotal"], r["tax"], r["total"],
                    r["paid_status"] if "paid_status" in r.keys() else "",
                    r["paid_date"] if "paid_date" in r.keys() else "",
                    r["due_date"],
                ])
            zf.writestr(f"{year}/invoices/_summary.csv", inv_csv.getvalue())
            for r in invoices:
                pdf_path = _paths["PDF_DIR"] / f"invoice_{r['invoice_number']:04d}.pdf"
                if pdf_path.exists():
                    zf.write(pdf_path, f"{year}/invoices/invoice_{r['invoice_number']:04d}.pdf")

            # Accounting docs
            bills = db.execute(
                "SELECT * FROM company_docs WHERE substr(doc_date,1,4)=? ORDER BY doc_date",
                (str(year),),
            ).fetchall()
            bills_csv = io.StringIO()
            w = csv.writer(bills_csv)
            w.writerow(["Date", "Vendor", "Description", "Amount", "Currency", "Category", "Status", "Due Date"])
            for r in bills:
                w.writerow([r["doc_date"], r["vendor"], r["description"],
                           r["amount"], r["currency"], r["category"], r["status"], r["due_date"] or ""])
            zf.writestr(f"{year}/accounting/_summary.csv", bills_csv.getvalue())
            for r in bills:
                if r["doc_file"]:
                    fp = _paths["ACCT_DIR"] / r["doc_file"]
                    if fp.exists():
                        ext = Path(r["doc_file"]).suffix
                        safe = r["vendor"].replace("/", "-")[:30]
                        zf.write(fp, f"{year}/accounting/docs/{r['doc_date']}_{safe}_{r['amount']}{ext}")

            # Travel expense reports — ALL of them (year-wide + each month-specific)
            try:
                rpt_keys = expense_reports[0].keys() if False else None
            except Exception:
                rpt_keys = None
            expense_reports_list = db.execute(
                "SELECT * FROM expense_reports WHERE year=? ORDER BY id",
                (year,),
            ).fetchall()
            for rep in expense_reports_list:
                rpt_num = rep["report_number"]
                # Detect whether month column exists (back-compat with older DBs)
                month_val = None
                try:
                    month_val = rep["month"]
                except (IndexError, KeyError):
                    month_val = None
                if month_val:
                    fname = f"expenses_{year}_{month_val:02d}_{rpt_num:04d}.pdf"
                    arc = f"{year}/travel_expenses/report_{year}-{month_val:02d}_{rpt_num:04d}.pdf"
                else:
                    fname = f"expenses_{year}_{rpt_num:04d}.pdf"
                    arc = f"{year}/travel_expenses/report_{rpt_num:04d}.pdf"
                fp = _paths["REPORT_DIR"] / fname
                if fp.exists():
                    zf.write(fp, arc)

            expenses = db.execute(
                "SELECT * FROM expenses WHERE substr(expense_date,1,4)=? ORDER BY expense_date",
                (str(year),),
            ).fetchall()
            exp_csv = io.StringIO()
            w = csv.writer(exp_csv)
            w.writerow(["Date", "Description", "Category", "Amount (CHF)", "Original Amount",
                        "Original Currency", "Scan filename"])
            for r in expenses:
                w.writerow([r["expense_date"], r["description"], r["category"],
                           r["amount"], r["original_amount"] or "", r["original_currency"] or "",
                           r["scan_file"] or ""])
            zf.writestr(f"{year}/travel_expenses/_summary.csv", exp_csv.getvalue())

            # ── Receipt scans for every expense (so Treuhand sees the originals) ──
            scan_dir = _paths.get("SCAN_DIR") or (_paths["REPORT_DIR"].parent / "scans")
            for r in expenses:
                if not r["scan_file"]:
                    continue
                fp = scan_dir / r["scan_file"]
                if not fp.exists():
                    continue
                ext = Path(r["scan_file"]).suffix
                safe = (r["description"] or "expense").replace("/", "-")[:40]
                arc_name = f"{year}/travel_expenses/scans/{r['expense_date']}_{safe}_{r['amount']}{ext}"
                zf.write(fp, arc_name)

            # ── Payslips (Lohnabrechnungen) for the year ──
            payslip_dir = _paths.get("PAYSLIP_DIR")
            payslips = db.execute(
                "SELECT id, year, month, gross, net_salary, total_employer_cost, "
                "pdf_file FROM payslips WHERE year=? ORDER BY month",
                (year,),
            ).fetchall() if _table_exists(db, "payslips") else []
            if payslips:
                ps_csv = io.StringIO()
                w = csv.writer(ps_csv)
                w.writerow(["Year", "Month", "Gross", "Net", "Employer cost", "PDF file"])
                for r in payslips:
                    w.writerow([r["year"], r["month"], r["gross"], r["net_salary"],
                                r["total_employer_cost"], r["pdf_file"] or ""])
                zf.writestr(f"{year}/payslips/_summary.csv", ps_csv.getvalue())
                if payslip_dir:
                    for r in payslips:
                        if r["pdf_file"]:
                            fp = payslip_dir / r["pdf_file"]
                            if fp.exists():
                                ext = Path(r["pdf_file"]).suffix
                                zf.write(fp, f"{year}/payslips/{year}-{r['month']:02d}_Lohnabrechnung{ext}")
            # Also include any raw payslip PDFs sitting in PAYSLIP_DIR (not in DB)
            if payslip_dir and payslip_dir.exists():
                for fp in sorted(payslip_dir.iterdir()):
                    if not fp.is_file():
                        continue
                    if fp.suffix.lower() not in (".pdf", ".jpg", ".jpeg", ".png"):
                        continue
                    # Skip if we already added it via the DB loop above
                    arc = f"{year}/payslips/{fp.name}"
                    if arc not in zf.namelist():
                        # Only include files that look like they belong to this year
                        if str(year) in fp.name:
                            zf.write(fp, arc)

            # ── Bank statements (UBS or other) for the year ──
            bank_dir = _paths.get("BANK_DIR")
            stmts = db.execute(
                "SELECT * FROM bank_statements WHERE substr(period_end,1,4)=? "
                "ORDER BY period_end",
                (str(year),),
            ).fetchall() if _table_exists(db, "bank_statements") else []
            if stmts:
                bs_csv = io.StringIO()
                w = csv.writer(bs_csv)
                w.writerow(["Period start", "Period end", "Bank", "Account", "IBAN",
                            "Type", "Currency", "Opening balance", "Closing balance",
                            "Notes", "PDF file", "XML file"])
                # Helper: detect both column names (post- and pre-migration)
                def _get_file(row, *keys):
                    for k in keys:
                        try:
                            v = row[k]
                            if v:
                                return v
                        except (IndexError, KeyError):
                            pass
                    return None
                for r in stmts:
                    pdf_name = _get_file(r, "statement_file_pdf", "statement_file")
                    xml_name = _get_file(r, "statement_file_xml")
                    w.writerow([
                        r["period_start"], r["period_end"], r["bank"],
                        r["account_label"] or "", r["iban"] or "",
                        r["statement_type"], r["currency"],
                        r["opening_balance"] if r["opening_balance"] is not None else "",
                        r["closing_balance"] if r["closing_balance"] is not None else "",
                        r["notes"] or "", pdf_name or "", xml_name or "",
                    ])
                zf.writestr(f"{year}/bank_statements/_summary.csv", bs_csv.getvalue())
                if bank_dir:
                    for r in stmts:
                        safe_acct = (r["account_label"] or r["bank"]).replace("/", "-")[:30]
                        for kind, attr in [("PDF", "statement_file_pdf"),
                                            ("XML", "statement_file_xml")]:
                            fname = _get_file(r, attr, "statement_file") if kind == "PDF" else _get_file(r, attr)
                            if not fname:
                                continue
                            fp = bank_dir / fname
                            if not fp.exists():
                                continue
                            ext = Path(fname).suffix
                            arc = f"{year}/bank_statements/{r['period_end']}_{safe_acct}{ext}"
                            # Avoid duplicate writes if same file was uploaded for both
                            if arc not in zf.namelist():
                                zf.write(fp, arc)

            # Obligations
            obs = db.execute(
                "SELECT * FROM obligations WHERE period_year=? ORDER BY due_date",
                (year,),
            ).fetchall()
            ob_csv = io.StringIO()
            w = csv.writer(ob_csv)
            w.writerow(["Type", "Period", "Amount", "Due Date", "Status", "Notes"])
            for r in obs:
                w.writerow([OBLIGATION_TYPES_REF().get(r["obligation_type"], r["obligation_type"]),
                           r["period_label"], r["amount"], r["due_date"] or "",
                           r["status"], r["notes"] or ""])
            zf.writestr(f"{year}/obligations/_summary.csv", ob_csv.getvalue())

    buf.seek(0)
    out_path = _paths["REPORT_DIR"] / f"accountant_package_{year}.zip"
    out_path.write_bytes(buf.getvalue())
    return FileResponse(
        out_path,
        filename=f"Muster Consulting Accountant Package {year}.zip",
        media_type="application/zip",
    )




# ─── Annual P&L + VAT + Tax (appended) ───────────────────────────────

@router.get("/pl/{year}")
async def pl_report(year: int):
    """Accrual-basis P&L for the year, matching what the tax return will see.

    Rules (each one fixes a former double-count):
    - Revenue = invoice SUBTOTALS (net of VAT — the 8.1% belongs to the ESTV,
      not to us), service invoices only (hours>0; hours=0 = travel pass-through).
    - Extra income = income_entries NOT linked to an invoice (paid invoices
      auto-create linked rows) and not 'Salary' (that's personal-side tracking).
    - Payroll cost = employer cost of the payslips actually ISSUED this year,
      not 12× the current setting.
    - Bills exclude 'Payroll Settlement' (AXA premiums already accrued in the
      payslip employer cost) and 'Taxes / VAT' (VAT payments are not a cost).
    - Obligations are NOT costs: AHV/BVG/UVG/KTG obligations are the payment
      side of the payroll cost above; VAT/corporate-tax obligations are not
      pre-tax expenses. Shown as an informational cash-owed breakdown only.
    """
    EXCLUDED_BILL_CATS = ("Payroll Settlement", "Taxes / VAT")
    with get_db() as db:
        # Revenue: net of VAT, service invoices only
        invoice_rev = db.execute(
            "SELECT COALESCE(SUM(subtotal),0) as t FROM invoices WHERE year=? AND hours>0",
            (year,),
        ).fetchone()["t"]
        invoice_paid = db.execute(
            "SELECT COALESCE(SUM(subtotal),0) as t FROM invoices WHERE year=? AND hours>0 AND paid_status='paid'",
            (year,),
        ).fetchone()["t"]
        extra_income = db.execute(
            """SELECT COALESCE(SUM(amount),0) as t FROM income_entries
               WHERE substr(income_date,1,4)=? AND invoice_id IS NULL AND category != 'Salary'""",
            (str(year),),
        ).fetchone()["t"]

        # Operating costs by category from company_docs
        cost_rows = db.execute(
            f"""SELECT category, COUNT(*) as cnt, SUM(amount) as total
               FROM company_docs WHERE substr(doc_date,1,4)=?
               AND category NOT IN ({','.join('?' * len(EXCLUDED_BILL_CATS))})
               GROUP BY category ORDER BY total DESC""",
            (str(year), *EXCLUDED_BILL_CATS),
        ).fetchall()

        # Travel expenses — REIMBURSABLE, kept separate from operating costs
        travel = db.execute(
            "SELECT COALESCE(SUM(amount),0) as t, COUNT(*) as cnt FROM expenses WHERE substr(expense_date,1,4)=?",
            (str(year),),
        ).fetchone()
        # Reimbursement = invoices with hours=0 (travel expense report invoices)
        travel_reimbursed = db.execute(
            "SELECT COALESCE(SUM(total),0) as t FROM invoices WHERE year=? AND hours=0",
            (year,),
        ).fetchone()["t"]

        # Obligations by type — informational (cash owed), NOT part of the P&L
        ob_rows = db.execute(
            """SELECT obligation_type, COUNT(*) as cnt, SUM(amount) as total,
                      SUM(CASE WHEN status='paid' THEN amount ELSE 0 END) as paid
               FROM obligations WHERE period_year=?
               GROUP BY obligation_type ORDER BY total DESC""",
            (year,),
        ).fetchall()

        # Payroll cost: employer cost of payslips actually issued this year
        salary_row = db.execute(
            "SELECT COALESCE(SUM(total_employer_cost),0) as t, COUNT(*) as n FROM payslips WHERE year=?",
            (year,),
        ).fetchone()
        salary_total, payslip_count = salary_row["t"], salary_row["n"]

    cost_categories = [
        {"category": r["category"], "count": r["cnt"], "total": r["total"]}
        for r in cost_rows
    ]
    travel_total = travel["t"]
    travel_count = travel["cnt"]
    ob_breakdown = [
        {
            "type": OBLIGATION_TYPES_REF().get(r["obligation_type"], r["obligation_type"]),
            "count": r["cnt"], "total": r["total"], "paid": r["paid"],
        }
        for r in ob_rows
    ]

    total_revenue = invoice_rev + extra_income
    total_costs = sum(c["total"] for c in cost_categories) + salary_total
    obligations_total = sum(o["total"] for o in ob_breakdown)
    profit_before_tax = round(total_revenue - total_costs, 2)

    # Travel net = how much the GmbH is currently out-of-pocket on travel
    travel_net = travel_total - travel_reimbursed

    return {
        "year": year,
        "basis": "Accrual, net of VAT. Payroll = issued payslips. Obligations shown for cash planning only (their P&L side already lives in payroll/bills).",
        "revenue": {
            "invoices_issued": invoice_rev,
            "invoices_paid": invoice_paid,
            "extra_income": extra_income,
            "total": total_revenue,
        },
        "costs": {
            "salary": salary_total,
            "payslip_count": payslip_count,
            "company_docs": cost_categories,
            "company_docs_total": sum(c["total"] for c in cost_categories),
            "excluded_categories": list(EXCLUDED_BILL_CATS),
            "total": total_costs,
        },
        "obligations": {
            "breakdown": ob_breakdown,
            "total": obligations_total,
            "note": "Cash owed to authorities/insurers — informational, not added to costs.",
        },
        "travel_pass_through": {
            "expenses_paid": travel_total,
            "expenses_count": travel_count,
            "reimbursed_by_client": travel_reimbursed,
            "net_outstanding": travel_net,
            "note": "Pass-through: not part of GmbH operating costs. Reimbursed via expense report invoices.",
        },
        "profit_before_tax": profit_before_tax,
        "profit_margin_pct": round((profit_before_tax / total_revenue * 100), 1) if total_revenue else 0,
    }


@router.get("/pl/{year}/excel")
async def pl_report_excel(year: int):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import io

    data = await pl_report(year)

    wb = Workbook()
    ws = wb.active
    ws.title = f"P&L {year}"

    bold = Font(bold=True, size=12)
    header_fill = PatternFill("solid", fgColor="D7E1E8")
    chf_fmt = "#,##0.00"

    row = 1
    ws.cell(row=row, column=1, value=f"Muster Consulting GmbH — Annual P&L {year}").font = Font(bold=True, size=16)
    row += 2

    # Revenue
    ws.cell(row=row, column=1, value="REVENUE").font = bold
    ws.cell(row=row, column=1).fill = header_fill
    row += 1
    ws.cell(row=row, column=1, value="Invoices issued")
    c = ws.cell(row=row, column=2, value=data["revenue"]["invoices_issued"]); c.number_format = chf_fmt
    row += 1
    ws.cell(row=row, column=1, value="  of which paid")
    c = ws.cell(row=row, column=2, value=data["revenue"]["invoices_paid"]); c.number_format = chf_fmt
    row += 1
    ws.cell(row=row, column=1, value="Other income")
    c = ws.cell(row=row, column=2, value=data["revenue"]["extra_income"]); c.number_format = chf_fmt
    row += 1
    ws.cell(row=row, column=1, value="Total Revenue").font = bold
    c = ws.cell(row=row, column=2, value=data["revenue"]["total"]); c.number_format = chf_fmt; c.font = bold
    row += 2

    # Costs
    ws.cell(row=row, column=1, value="COSTS").font = bold
    ws.cell(row=row, column=1).fill = header_fill
    row += 1
    ws.cell(row=row, column=1, value="Salary (annual)")
    c = ws.cell(row=row, column=2, value=data["costs"]["salary"]); c.number_format = chf_fmt
    row += 1
    for cat in data["costs"]["company_docs"]:
        ws.cell(row=row, column=1, value=f"  {cat['category']}")
        c = ws.cell(row=row, column=2, value=cat["total"]); c.number_format = chf_fmt
        row += 1
    ws.cell(row=row, column=1, value="Total Costs").font = bold
    c = ws.cell(row=row, column=2, value=data["costs"]["total"]); c.number_format = chf_fmt; c.font = bold
    row += 2

    # Obligations
    ws.cell(row=row, column=1, value="OBLIGATIONS (AHV / BVG / Tax)").font = bold
    ws.cell(row=row, column=1).fill = header_fill
    row += 1
    for ob in data["obligations"]["breakdown"]:
        ws.cell(row=row, column=1, value=f"  {ob['type']}")
        c = ws.cell(row=row, column=2, value=ob["total"]); c.number_format = chf_fmt
        row += 1
    ws.cell(row=row, column=1, value="Total Obligations").font = bold
    c = ws.cell(row=row, column=2, value=data["obligations"]["total"]); c.number_format = chf_fmt; c.font = bold
    row += 2

    # Profit
    ws.cell(row=row, column=1, value="PROFIT BEFORE TAX").font = Font(bold=True, size=14)
    c = ws.cell(row=row, column=2, value=data["profit_before_tax"]); c.number_format = chf_fmt; c.font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value="Profit margin %")
    c = ws.cell(row=row, column=2, value=f"{data['profit_margin_pct']}%")
    row += 2

    # Travel pass-through (separate from operating P&L)
    tp = data["travel_pass_through"]
    ws.cell(row=row, column=1, value="TRAVEL (PASS-THROUGH — NOT IN P&L)").font = bold
    ws.cell(row=row, column=1).fill = header_fill
    row += 1
    ws.cell(row=row, column=1, value=f"  Expenses paid ({tp['expenses_count']} receipts)")
    c = ws.cell(row=row, column=2, value=tp["expenses_paid"]); c.number_format = chf_fmt
    row += 1
    ws.cell(row=row, column=1, value="  Reimbursed by client")
    c = ws.cell(row=row, column=2, value=tp["reimbursed_by_client"]); c.number_format = chf_fmt
    row += 1
    ws.cell(row=row, column=1, value="  Net outstanding (still owed by client)").font = bold
    c = ws.cell(row=row, column=2, value=tp["net_outstanding"]); c.number_format = chf_fmt; c.font = bold

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fp = _paths["REPORT_DIR"] / f"PL_{year}.xlsx"
    fp.write_bytes(buf.getvalue())
    return FileResponse(
        fp,
        filename=f"Muster Consulting P&L {year}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _load_vat_settings(db) -> dict:
    import json as _json
    row = db.execute("SELECT * FROM vat_settings WHERE id=1").fetchone()
    if not row:
        return {"estimate_missing": True, "estimate_rate": 8.1,
                "excluded_categories": ["Insurance", "Bank Fees", "Payroll Settlement"],
                "flat_quarterly_deduction": 0.0}
    try:
        excluded = _json.loads(row["excluded_categories"])
    except Exception:
        excluded = []
    return {
        "estimate_missing": bool(row["estimate_missing"]),
        "estimate_rate": float(row["estimate_rate"]),
        "excluded_categories": excluded,
        "flat_quarterly_deduction": float(row["flat_quarterly_deduction"]),
    }


@tax_router.get("/vat/settings")
async def get_vat_settings():
    with get_db() as db:
        return _load_vat_settings(db)


@tax_router.put("/vat/settings")
async def update_vat_settings(body: dict):
    """Partial update: fields absent from the body keep their current value,
    so a stray empty PUT can never reset the settings."""
    import json as _json
    with get_db() as db:
        current = _load_vat_settings(db)
        estimate_missing = 1 if body.get("estimate_missing", current["estimate_missing"]) else 0
        estimate_rate = float(body.get("estimate_rate", current["estimate_rate"]))
        if not 0 <= estimate_rate <= 100:
            raise HTTPException(400, "estimate_rate must be between 0 and 100")
        excluded = body.get("excluded_categories", current["excluded_categories"])
        if not isinstance(excluded, list):
            raise HTTPException(400, "excluded_categories must be a list")
        flat = float(body.get("flat_quarterly_deduction", current["flat_quarterly_deduction"]))
        if flat < 0:
            raise HTTPException(400, "flat_quarterly_deduction cannot be negative")
        db.execute(
            """UPDATE vat_settings SET estimate_missing=?, estimate_rate=?,
               excluded_categories=?, flat_quarterly_deduction=?, updated_at=datetime('now')
               WHERE id=1""",
            (estimate_missing, estimate_rate, _json.dumps(excluded), flat),
        )
        return _load_vat_settings(db)


def _vat_quarter_data(db, year: int, quarter: int, settings: dict) -> dict:
    """Effective-method VAT for one quarter: output VAT from invoices, input
    VAT from recorded bill VAT + simulated deductions per vat_settings."""
    start_month = (quarter - 1) * 3 + 1
    end_month = start_month + 2

    output_vat = db.execute(
        "SELECT COALESCE(SUM(tax),0) as t FROM invoices WHERE year=? AND month>=? AND month<=? AND hours>0",
        (year, start_month, end_month),
    ).fetchone()["t"]

    # Input VAT explicitly recorded on bills
    input_recorded = db.execute(
        """SELECT COALESCE(SUM(vat_amount),0) as t FROM company_docs
           WHERE substr(doc_date,1,4)=? AND CAST(substr(doc_date,6,2) AS INTEGER) BETWEEN ? AND ?
           AND vat_amount > 0""",
        (str(year), start_month, end_month),
    ).fetchone()["t"]

    # Simulated: bills without a recorded VAT amount are assumed to include
    # VAT at estimate_rate (deductible share = amount × r / (100 + r)),
    # except VAT-exempt categories.
    input_estimated = 0.0
    estimated_bills = 0
    if settings["estimate_missing"]:
        excluded = set(settings["excluded_categories"])
        r = settings["estimate_rate"]
        rows = db.execute(
            """SELECT category, COALESCE(SUM(amount),0) as t, COUNT(*) as n FROM company_docs
               WHERE substr(doc_date,1,4)=? AND CAST(substr(doc_date,6,2) AS INTEGER) BETWEEN ? AND ?
               AND (vat_amount IS NULL OR vat_amount = 0) AND amount > 0
               GROUP BY category""",
            (str(year), start_month, end_month),
        ).fetchall()
        for row in rows:
            if row["category"] in excluded:
                continue
            input_estimated += row["t"] * r / (100 + r)
            estimated_bills += row["n"]
    input_estimated = round(input_estimated, 2)

    flat = settings["flat_quarterly_deduction"]
    total_deductions = round(input_recorded + input_estimated + flat, 2)
    vat_due = round(output_vat - total_deductions, 2)

    # Swiss filing rule: declare + pay within 60 days of quarter end
    q_end = date(year, end_month, calendar.monthrange(year, end_month)[1])
    due_date = date.fromordinal(q_end.toordinal() + 60).isoformat()

    obligation = db.execute(
        "SELECT id, amount, status, due_date, doc_file FROM obligations WHERE obligation_type='vat' AND period_label=?",
        (f"Q{quarter} {year}",),
    ).fetchone()

    return {
        "year": year,
        "quarter": quarter,
        "period": f"Q{quarter} {year}",
        "output_vat": round(output_vat, 2),
        "input_vat_recorded": round(input_recorded, 2),
        "input_vat_estimated": input_estimated,
        "estimated_bills": estimated_bills,
        "flat_deduction": round(flat, 2),
        "input_vat": total_deductions,        # backward-compatible: total deductions
        "vat_due": vat_due,
        "due_date": due_date,
        "obligation": dict(obligation) if obligation else None,
    }


@tax_router.get("/vat/{year}/{quarter}")
async def vat_quarter(year: int, quarter: int):
    """Q1=1-3, Q2=4-6, Q3=7-9, Q4=10-12"""
    if quarter < 1 or quarter > 4:
        raise HTTPException(400, "Quarter must be 1-4")
    with get_db() as db:
        settings = _load_vat_settings(db)
        return _vat_quarter_data(db, year, quarter, settings)


@tax_router.post("/vat/{year}/{quarter}/obligation")
async def create_vat_obligation(year: int, quarter: int):
    """Create (or refresh) the quarterly VAT obligation so it shows up in
    Obligations and the Calendar. Refuses to touch a paid obligation."""
    if quarter < 1 or quarter > 4:
        raise HTTPException(400, "Quarter must be 1-4")
    with get_db() as db:
        settings = _load_vat_settings(db)
        data = _vat_quarter_data(db, year, quarter, settings)
        if data["vat_due"] <= 0:
            raise HTTPException(400, f"No VAT due for Q{quarter} {year} "
                                     f"(net position: {data['vat_due']:.2f})")
        label = f"Q{quarter} {year}"
        notes = (f"VAT {label} — output {data['output_vat']:.2f}, deductions "
                 f"{data['input_vat']:.2f} (recorded {data['input_vat_recorded']:.2f}, "
                 f"estimated {data['input_vat_estimated']:.2f}, flat {data['flat_deduction']:.2f})")
        existing = data["obligation"]
        if existing:
            if existing["status"] == "paid":
                raise HTTPException(400, f"VAT obligation for {label} is already paid")
            if existing["doc_file"]:
                raise HTTPException(400,
                    f"VAT obligation for {label} was readjusted from an uploaded assessment — "
                    "its amount is authoritative. Edit it on the Obligations page if needed.")
            db.execute(
                "UPDATE obligations SET amount=?, due_date=?, notes=? WHERE id=?",
                (data["vat_due"], data["due_date"], notes, existing["id"]),
            )
            return {"id": existing["id"], "updated": True, "amount": data["vat_due"]}
        cur = db.execute(
            """INSERT INTO obligations
               (obligation_type, period_label, period_year, amount, currency,
                due_date, status, notes, recurrence)
               VALUES ('vat', ?, ?, ?, 'CHF', ?, 'unpaid', ?, 'none')""",
            (label, year, data["vat_due"], data["due_date"], notes),
        )
        return {"id": cur.lastrowid, "updated": False, "amount": data["vat_due"]}


@tax_router.get("/tax/estimate/{year}")
async def estimate_tax(year: int):
    """Rough Swiss corporate tax estimate."""
    pl = await pl_report(year)
    profit = pl["profit_before_tax"]
    # Swiss federal corporate tax ~8.5% on profit
    federal = max(0, profit * 0.085)
    # Cantonal/communal varies wildly (e.g. Zurich ~13%)
    cantonal = max(0, profit * 0.13)
    return {
        "year": year,
        "profit_before_tax": profit,
        "federal_tax_estimate": round(federal, 2),
        "cantonal_tax_estimate": round(cantonal, 2),
        "total_tax_estimate": round(federal + cantonal, 2),
        "effective_rate_pct": round((federal + cantonal) / profit * 100, 1) if profit > 0 else 0,
        "note": "Estimate uses 8.5% federal + 13% cantonal (Zurich). Actual rate depends on canton.",
    }


